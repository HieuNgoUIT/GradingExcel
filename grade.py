from openpyxl import Workbook
import openpyxl
import os
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

def read_answer_from_txt(file):
    answers = []
    with open(file) as f:
        lines = f.read()
        parts = lines.split("-")
        for p in parts:
            value = p.strip()
            answers.append(value.split("\n"))
    return answers

def read_user_value_from_excel(ws, cell, start, end):
    users_answers = []
    for i in range(start,end):
        try:
            value = ws[cell + str(i)].value.strip()
        except:
            value = ""
            print("something happened at", cell, i, file)
        #assert isinstance(value, str)
        #assert len(value) == 1 or len(value) == 0
        users_answers.append(value)
    return users_answers

def write_answers(ws, teacher_column, start,end, answers):
    status = 0
    ws[teacher_column + str(start-1)] = 'Đáp án GV'
    for i in range(start,end):
        ws[teacher_column + str(i)] = answers[status]
        status += 1

def grade_per_file(ws, answers, student_column, total_column, row_start, row_end):
    sum_row = row_start-1
    count = row_start

    user_answers = read_user_value_from_excel(ws,student_column, row_start, row_end)
    sum = 0
    for user, answer in zip(user_answers, answers):
        if user == answer:
            ws[total_column+ str(count)] = 1
            ws[total_column+ str(count)].alignment = Alignment(horizontal="center")
            sum += 1
        else:
            ws[total_column+ str(count)] = 0
            ws[total_column+ str(count)].alignment = Alignment(horizontal="center")
        count += 1
    ws[total_column + str(sum_row)] = sum 
    ws[total_column + str(sum_row)].alignment = Alignment(horizontal="center")

if __name__ == "__main__": 
    pwd = os.getcwd()
    part_ans = read_answer_from_txt(os.path.join(pwd, "answer.txt"))
    files = os.listdir(os.path.join(pwd, "hocsinh"))
    for file in files:
        wb = openpyxl.load_workbook(os.path.join(pwd, "hocsinh", file))
        ws = wb.worksheets[0]
		
        
        write_answers(ws, "C", 9, 15, part_ans[0]) #ghi dap an vao` cot. C
        write_answers(ws, "H", 9, 34, part_ans[1])
        write_answers(ws, "M", 9, 21, part_ans[2])
        write_answers(ws, "R", 9, 21, part_ans[3])
        write_answers(ws, "W", 9, 39, part_ans[4])
        write_answers(ws, "AB", 9, 13, part_ans[5])
        write_answers(ws, "AG", 9, 19, part_ans[6])

        grade_per_file(ws, part_ans[0], student_column="B", total_column="D", row_start=9 , row_end=15)
        grade_per_file(ws, part_ans[1], student_column="G", total_column="I", row_start=9 , row_end=34)
        grade_per_file(ws, part_ans[2], student_column="L", total_column="N", row_start=9 , row_end=21)
        grade_per_file(ws, part_ans[3], student_column="Q", total_column="S", row_start=9 , row_end=21) # row_end should +1 due to excel
        grade_per_file(ws, part_ans[4],  student_column="V", total_column="X", row_start=9 , row_end=39)
        grade_per_file(ws, part_ans[5],  student_column="AA", total_column="AC", row_start=9 , row_end=13)
        grade_per_file(ws, part_ans[6],  student_column="AF", total_column="AH", row_start=9 , row_end=19)

        wb.save(os.path.join(pwd, "result", file))
